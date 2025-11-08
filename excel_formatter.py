#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Data Formatter - Программа для форматирования Excel данных
Поддерживает различные операции форматирования для Windows
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class ExcelFormatter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Formatter")
        self.root.geometry("800x600")

        # Переменные
        self.file_path = tk.StringVar()
        self.status_text = tk.StringVar(value="Готов к работе")

        self.create_widgets()

    def create_widgets(self):
        # Фрейм для выбора файла
        file_frame = ttk.LabelFrame(self.root, text="Выбор файла", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Entry(file_frame, textvariable=self.file_path, width=60).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="Обзор...", command=self.browse_file).pack(side=tk.LEFT)

        # Фрейм с операциями форматирования
        operations_frame = ttk.LabelFrame(self.root, text="Операции форматирования", padding=10)
        operations_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Левая колонка - основное форматирование
        left_frame = ttk.Frame(operations_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        ttk.Label(left_frame, text="Основное форматирование:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=5)

        ttk.Button(left_frame, text="Форматировать заголовки",
                  command=self.format_headers, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(left_frame, text="Автоподбор ширины колонок",
                  command=self.auto_fit_columns, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(left_frame, text="Добавить границы",
                  command=self.add_borders, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(left_frame, text="Форматировать числа",
                  command=self.format_numbers, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(left_frame, text="Центрировать текст",
                  command=self.center_align, width=30).pack(pady=3, fill=tk.X)

        # Правая колонка - дополнительные функции
        right_frame = ttk.Frame(operations_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        ttk.Label(right_frame, text="Дополнительно:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=5)

        ttk.Button(right_frame, text="Удалить пустые строки",
                  command=self.remove_empty_rows, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(right_frame, text="Удалить дубликаты",
                  command=self.remove_duplicates, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(right_frame, text="Заморозить первую строку",
                  command=self.freeze_panes, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(right_frame, text="Сортировать по первой колонке",
                  command=self.sort_by_first_column, width=30).pack(pady=3, fill=tk.X)
        ttk.Button(right_frame, text="Полное форматирование",
                  command=self.full_format, width=30).pack(pady=3, fill=tk.X)

        # Фрейм статуса
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(status_frame, text="Статус:").pack(side=tk.LEFT)
        ttk.Label(status_frame, textvariable=self.status_text,
                 foreground='green').pack(side=tk.LEFT, padx=5)

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            self.status_text.set(f"Выбран файл: {os.path.basename(filename)}")

    def check_file(self):
        if not self.file_path.get():
            messagebox.showerror("Ошибка", "Пожалуйста, выберите файл Excel")
            return False
        if not os.path.exists(self.file_path.get()):
            messagebox.showerror("Ошибка", "Файл не найден")
            return False
        return True

    def format_headers(self):
        """Форматирование заголовков (первая строка)"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active

            # Стиль для заголовков
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Заголовки отформатированы успешно")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка форматирования: {str(e)}")

    def auto_fit_columns(self):
        """Автоматическая подгонка ширины колонок"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Ширина колонок подогнана автоматически")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка подгонки колонок: {str(e)}")

    def add_borders(self):
        """Добавление границ ко всем ячейкам с данными"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.border = thin_border

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Границы добавлены успешно")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка добавления границ: {str(e)}")

    def format_numbers(self):
        """Форматирование числовых данных"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active

            for row in ws.iter_rows(min_row=2):  # Пропускаем заголовки
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Числа отформатированы успешно")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка форматирования чисел: {str(e)}")

    def center_align(self):
        """Центрирование текста во всех ячейках"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active

            center_alignment = Alignment(horizontal="center", vertical="center")

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.alignment = center_alignment

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Текст центрирован успешно")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка центрирования: {str(e)}")

    def remove_empty_rows(self):
        """Удаление пустых строк"""
        if not self.check_file():
            return

        try:
            df = pd.read_excel(self.file_path.get())
            df_cleaned = df.dropna(how='all')

            output_path = self.get_output_path()
            df_cleaned.to_excel(output_path, index=False)

            removed = len(df) - len(df_cleaned)
            self.status_text.set(f"Удалено пустых строк: {removed}")
            messagebox.showinfo("Успех", f"Удалено {removed} пустых строк. Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка удаления пустых строк: {str(e)}")

    def remove_duplicates(self):
        """Удаление дубликатов"""
        if not self.check_file():
            return

        try:
            df = pd.read_excel(self.file_path.get())
            initial_count = len(df)
            df_unique = df.drop_duplicates()

            output_path = self.get_output_path()
            df_unique.to_excel(output_path, index=False)

            removed = initial_count - len(df_unique)
            self.status_text.set(f"Удалено дубликатов: {removed}")
            messagebox.showinfo("Успех", f"Удалено {removed} дубликатов. Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка удаления дубликатов: {str(e)}")

    def freeze_panes(self):
        """Заморозка первой строки"""
        if not self.check_file():
            return

        try:
            wb = load_workbook(self.file_path.get())
            ws = wb.active
            ws.freeze_panes = 'A2'

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            self.status_text.set("Первая строка заморожена")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка заморозки панелей: {str(e)}")

    def sort_by_first_column(self):
        """Сортировка по первой колонке"""
        if not self.check_file():
            return

        try:
            df = pd.read_excel(self.file_path.get())
            first_column = df.columns[0]
            df_sorted = df.sort_values(by=first_column)

            output_path = self.get_output_path()
            df_sorted.to_excel(output_path, index=False)

            self.status_text.set("Данные отсортированы по первой колонке")
            messagebox.showinfo("Успех", f"Файл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка сортировки: {str(e)}")

    def full_format(self):
        """Полное форматирование (все операции)"""
        if not self.check_file():
            return

        try:
            # Читаем и очищаем данные
            df = pd.read_excel(self.file_path.get())
            df = df.dropna(how='all')  # Удаляем пустые строки
            df = df.drop_duplicates()  # Удаляем дубликаты

            # Сохраняем временно
            temp_path = self.file_path.get().replace('.xlsx', '_temp.xlsx')
            df.to_excel(temp_path, index=False)

            # Применяем форматирование
            wb = load_workbook(temp_path)
            ws = wb.active

            # Форматирование заголовков
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Границы
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Применяем стили к данным
            center_alignment = Alignment(horizontal="center", vertical="center")

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.border = thin_border
                        if cell.row > 1:  # Не для заголовков
                            cell.alignment = center_alignment
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'

            # Автоподбор ширины
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Заморозка панелей
            ws.freeze_panes = 'A2'

            output_path = self.get_output_path()
            wb.save(output_path)
            wb.close()

            # Удаляем временный файл
            if os.path.exists(temp_path):
                os.remove(temp_path)

            self.status_text.set("Полное форматирование выполнено успешно")
            messagebox.showinfo("Успех", f"Полное форматирование завершено!\nФайл сохранен: {output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка полного форматирования: {str(e)}")

    def get_output_path(self):
        """Создает путь для выходного файла"""
        original_path = self.file_path.get()
        base, ext = os.path.splitext(original_path)
        return f"{base}_formatted{ext}"


def main():
    root = tk.Tk()
    app = ExcelFormatter(root)
    root.mainloop()


if __name__ == "__main__":
    main()
